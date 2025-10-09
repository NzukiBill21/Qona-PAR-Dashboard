# routes/export.py
from __future__ import annotations

from flask import Blueprint, send_file, jsonify, request
from io import BytesIO, StringIO
from datetime import datetime
import logging
import re

import pandas as pd

from adapters.excel_loader import load_excel_raw
from services.data_service import (
    get_overall_and_weeks,
    get_officer_series,
    filter_payload_by_week,
)

bp = Blueprint("export", __name__)
log = logging.getLogger("excel")


# -----------------------------
# Optional drawing libraries
# -----------------------------
try:
    import matplotlib
    matplotlib.use("Agg")  # headless
    import matplotlib.pyplot as plt
    HAVE_MPL = True
except Exception:
    HAVE_MPL = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet
    HAVE_RL = True
except Exception:
    HAVE_RL = False


# =============================
# Helpers
# =============================
def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (s or "").strip().lower())

def _as_payload(officer: str | None, week: str | None, mode: str = "single") -> dict:
    """
    Build the same structure the UI uses:
      { snapshot: {...}, weekly: [...], officers: [...] }
    and apply week filter if provided.
    """
    data = get_officer_series(officer) if officer else get_overall_and_weeks()
    if week:
        # always prefer exact-date alignment unless caller overrides
        try:
            data = filter_payload_by_week(data, week, mode)
        except TypeError:
            # if older signature (data, week) exists in your codebase
            data = filter_payload_by_week(data, week)  # type: ignore
    return data

def _weekly_df(payload: dict) -> pd.DataFrame:
    weekly = payload.get("weekly", []) or []
    cols = ["week", "par", "outstanding", "provision"]
    df = pd.DataFrame(weekly, columns=cols)
    # keep order stable by parsing week if it's ISO or "5 Sep" etc
    def _to_date(s: str):
        for fmt in ("%Y-%m-%d", "%d %b", "%d %b %Y"):
            try:
                # If no year, use current year to sort
                val = datetime.strptime(s, fmt)
                if fmt == "%d %b":
                    val = val.replace(year=datetime.now().year)
                return val
            except Exception:
                continue
        return None
    if not df.empty:
        df["_ord"] = df["week"].map(_to_date)
        df = df.sort_values(by="_ord", kind="mergesort").drop(columns=["_ord"])
    return df

def _raw_df() -> pd.DataFrame:
    df = load_excel_raw()
    if df is None or df.empty:
        raise ValueError("No data")
    return df


# =============================
#            ROUTE
# =============================
@bp.get("/export")
def export_any():
    """
    /api/export?format=excel|xlsx|csv|json|png|pdf
    Optional: &officer=<name>&week=YYYY-MM-DD&mode=single|upto
    """
    fmt = (request.args.get("format") or "excel").lower()
    if fmt == "xlsx":  # alias
        fmt = "excel"

    officer = (
        request.args.get("officer")
        or request.args.get("name")
        or ""
    ).strip()
    week = request.args.get("week")
    mode = (request.args.get("mode") or "single").strip().lower()
    officer_disp = officer or "Overall Portfolio"
    officer_slug = _slug(officer) or "overall"

    log.info(f"[EXPORT] fmt={fmt} officer='{officer or 'ALL'}' week={week} mode={mode}")

    # Structured payload used by UI (snapshot + weekly)
    payload = _as_payload(officer, week, mode)

    # Raw DF (for CSV/Excel raw dump if needed)
    try:
        df_raw = _raw_df()
    except Exception:
        df_raw = pd.DataFrame()

    # ============== JSON ==============
    if fmt == "json":
        if df_raw.empty:
            return jsonify({"error": "No data"}), 404
        return jsonify(df_raw.to_dict(orient="records"))

    # ============== CSV (tidy weekly) ==============
    if fmt == "csv":
        df = _weekly_df(payload)
        if df.empty and not df_raw.empty:
            # fallback to raw if weekly missing
            df = df_raw
        buf = StringIO()
        df.to_csv(buf, index=False)
        buf.seek(0)
        name = f"par_{officer_slug}_{week or datetime.now().date()}.csv"
        return send_file(BytesIO(buf.getvalue().encode("utf-8")),
                         as_attachment=True,
                         download_name=name,
                         mimetype="text/csv")

    # ============== EXCEL (professional layout) ==============
    if fmt == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, NamedStyle
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ---- Sheet 1: Snapshot ----
        ws1 = wb.active
        ws1.title = "Snapshot"

        title = f"Qona Financial Services — PAR Analytics Report ({officer_disp})"
        ws1.merge_cells("A1:D1")
        c = ws1["A1"]
        c.value = title
        c.font = Font(size=14, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor="1E40AF")

        snap = payload.get("snapshot", {}) or {}
        kpi_rows = [
            ["Week", str(snap.get("week") or "")],
            ["Total Outstanding", snap.get("totalOutstanding", 0)],
            ["Total Gross Provision", snap.get("totalGrossProvision", 0)],
            ["Overall PAR (%)", snap.get("overallPAR", 0)],
        ]
        ws1.append([])  # row 2 empty
        ws1.append(["Metric", "Value"])  # row 3
        for r in kpi_rows:
            ws1.append(r)

        # style header row on Snapshot
        hdr = ws1[3]
        for cell in hdr:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.alignment = Alignment(horizontal="center")

        # number formats
        money = NamedStyle(name="money")
        money.number_format = "#,##0.00"
        pct = NamedStyle(name="pct")
        pct.number_format = "0.00"

        if "money" not in wb.named_styles:
            wb.add_named_style(money)
        if "pct" not in wb.named_styles:
            wb.add_named_style(pct)

        ws1["B5"].style = "money"
        ws1["B6"].style = "money"
        ws1["B7"].style = "pct"

        # widths
        ws1.column_dimensions["A"].width = 24
        ws1.column_dimensions["B"].width = 26
        ws1.freeze_panes = "A4"

        # ---- Sheet 2: Weekly ----
        ws2 = wb.create_sheet("Weekly")
        ws2.append(["Week", "PAR (%)", "Outstanding", "Provision"])

        dfw = _weekly_df(payload)
        for _, r in dfw.iterrows():
            ws2.append([r["week"], r["par"], r["outstanding"], r["provision"]])

        # header style
        thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin

        # col widths + number formats
        max_row = ws2.max_row
        for col_idx, key in enumerate(["Week", "PAR (%)", "Outstanding", "Provision"], start=1):
            letter = get_column_letter(col_idx)
            # auto width
            width = max(len(str(ws2[letter + "1"].value)), 10)
            for r in range(2, max_row + 1):
                val = ws2[f"{letter}{r}"].value
                width = max(width, len(f"{val:,}") if isinstance(val, (int, float)) else len(str(val)))
            ws2.column_dimensions[letter].width = min(width + 2, 40)

        # formats
        for r in ws2.iter_rows(min_row=2, min_col=2, max_col=3, max_row=max_row):
            # col B (PAR) -> pct style but we keep 4.19 not 0.0419
            r[0].number_format = "0.00"
            # col C (Outstanding)
            r[1].number_format = "#,##0.00"
        for r in ws2.iter_rows(min_row=2, min_col=4, max_col=4, max_row=max_row):
            r[0].number_format = "#,##0.00"

        ws2.freeze_panes = "A2"

        # ---- (Optional) Sheet 3: Raw (if available) ----
        if not df_raw.empty:
            ws3 = wb.create_sheet("Raw")
            ws3.append(list(df_raw.columns))
            for _, row in df_raw.iterrows():
                ws3.append([row.get(col) for col in df_raw.columns])
            # simple width
            for i, col in enumerate(df_raw.columns, start=1):
                ws3.column_dimensions[get_column_letter(i)].width = min(max(len(col), 14), 40)
            ws3.freeze_panes = "A2"

        # save to bytes
        out = BytesIO()
        wb.save(out)
        out.seek(0)

        name = f"par_report_{officer_slug}_{week or datetime.now().date()}.xlsx"
        return send_file(
            out,
            as_attachment=True,
            download_name=name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ============== PNG (Officer-aware trend) ==============
    if fmt == "png":
        if not HAVE_MPL:
            return jsonify({"error": "PNG export requires matplotlib"}), 501

        weekly = payload.get("weekly", []) or []
        if not weekly:
            return jsonify({"error": "No weekly data to export"}), 404

        x = [w["week"] for w in weekly]
        y = [w["par"] for w in weekly]

        # Make it look polished
        fig, ax = plt.subplots(figsize=(10.5, 4.5))
        ax.plot(x, y, marker="o", linewidth=2.25, color="#2563EB", label="PAR (%)")
        ax.fill_between(range(len(y)), y, [min(y)] * len(y), alpha=0.08, color="#2563EB")
        ax.grid(alpha=0.25, linestyle="--", linewidth=0.6)
        ax.set_title(f"PAR Trend — {officer_disp}", fontsize=13, fontweight="bold")
        ax.set_xlabel("Week")
        ax.set_ylabel("PAR (%)")
        # annotate last point
        ax.annotate(f"{y[-1]:.2f}%", xy=(len(y) - 1, y[-1]),
                    xytext=(len(y) - 1, y[-1] + max(y) * 0.05),
                    ha="center", arrowprops=dict(arrowstyle="->", lw=0.8))
        fig.tight_layout()

        buf = BytesIO()
        plt.savefig(buf, format="png", dpi=220, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)

        name = f"par_trend_{officer_slug}_{datetime.now().strftime('%Y%m%d')}.png"
        return send_file(buf, as_attachment=True, download_name=name, mimetype="image/png")

    # ============== PDF (clean, week-aligned) ==============
    if fmt == "pdf":
        if not HAVE_RL:
            return jsonify({"error": "PDF export requires reportlab"}), 501

        out = BytesIO()
        doc = SimpleDocTemplate(
            out,
            pagesize=A4,
            title=f"Portfolio Analytics Report — {officer_disp}",
        )
        styles = getSampleStyleSheet()
        story = []

        week_label = payload.get("snapshot", {}).get("week", "") or ""
        story.append(
            Paragraph(
                f"<b>Portfolio Analytics Report — {officer_disp}</b>"
                f"<br/><font size=10 color='grey'>Week: {week_label}</font>",
                styles["Title"],
            )
        )
        story.append(Spacer(1, 10))

        snap = payload.get("snapshot", {}) or {}
        kpi_rows = [
            ["Week", str(week_label)],
            ["Total Outstanding", f"{snap.get('totalOutstanding', 0):,.2f}"],
            ["Total Gross Provision", f"{snap.get('totalGrossProvision', 0):,.2f}"],
            ["Overall PAR (%)", f"{snap.get('overallPAR', 0):,.2f}"],
        ]
        kpi_table = Table(kpi_rows, colWidths=[180, 240])
        kpi_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                    ("FONT", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                ]
            )
        )
        story.append(kpi_table)
        story.append(Spacer(1, 12))

        weekly = payload.get("weekly", []) or []
        story.append(Paragraph("Weekly Breakdown", styles["Heading3"]))
        if weekly:
            tbl_rows = [["Week", "PAR (%)", "Outstanding", "Provision"]]
            for w in weekly:
                tbl_rows.append([
                    w["week"],
                    f"{w['par']:,.2f}",
                    f"{w['outstanding']:,.2f}",
                    f"{w['provision']:,.2f}",
                ])
            tbl = Table(tbl_rows, colWidths=[110, 70, 150, 150])
            tbl.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                        ("FONT", (0, 0), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ]
                )
            )
            story.append(tbl)
        else:
            story.append(Paragraph("No weekly data available for the selected period.", styles["Normal"]))

        doc.build(story)
        out.seek(0)
        name = f"par_report_{officer_slug}_{week or datetime.now().date()}.pdf"
        return send_file(out, as_attachment=True, download_name=name, mimetype="application/pdf")

    # ============== Unsupported ==============
    return jsonify({"error": f"Unsupported format '{fmt}'"}), 400
